from selenium import webdriver
import time

import openpyxl

from src.scrapy.journeyinlife.journey.journey.locator import bluesprint_locator
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from src.scrapy.journeyinlife.journey.journey.items import JourneyItem
import xlsxwriter
import hashlib
import db_cassandra

time_loading = 2


class JourneyParse:
    page_total = 990
    driver = None
    page_number = 975
    count = 11689
    workbook = None
    worksheet = None

    def init_excel(self):
        try:
            self.workbook = openpyxl.load_workbook(f'journey_in_life_{self.page_total}.xlsx')
        except FileNotFoundError:
            openpyxl.Workbook().save(f'journey_in_life_{self.page_total}.xlsx')
            self.workbook = openpyxl.load_workbook(f'journey_in_life_{self.page_total}.xlsx')

        self.worksheet = self.workbook.active

        cell_obj = self.worksheet.cell(row=1, column=1)
        if cell_obj.value is None:
            self.worksheet['A1'].value = 'STT'
            self.worksheet['B1'].value = 'Key'
            self.worksheet['C1'].value = 'Title'
            self.worksheet['D1'].value = 'Meaning'
            self.worksheet['E1'].value = 'Link'
            self.worksheet['F1'].value = 'Image_Link'
            self.worksheet['G1'].value = 'thumb_Link'
            self.__save_ex()

        print(str(self.worksheet.max_row))

    def __save_ex(self):
        self.workbook.save(f'journey_in_life_{self.page_total}.xlsx')

    def init_selenium(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--incognito')
        options.add_argument('--headless')
        self.driver = webdriver.Chrome('./chromedriver', chrome_options=options)

    def get_data_excel(self):
        for i in range(1378, 2400):  # range(self.worksheet.max_row):

            start_time = time.time()
            row_number = i + 2
            print(self.worksheet[f'A{row_number}'].value)
            print(self.worksheet[f'C{row_number}'].value)
            self.parse_detail_url(self.worksheet[f'E{row_number}'].value, row_number)
            print("--- %s seconds ---" % (time.time() - start_time))

    def get_data_excel_to_list(self):
        list_jr = []
        for i in range(1, self.worksheet.max_row):
            row_number = i + 2
            print(self.worksheet[f'A{row_number}'].value)
            print(self.worksheet[f'C{row_number}'].value)
            item_ = JourneyItem()
            item_.key = self.worksheet[f'B{row_number}'].value
            item_.title = self.worksheet[f'B{row_number}'].value
            item_.link = self.worksheet[f'E{row_number}'].value
            item_.thumb = self.worksheet[f'G{row_number}'].value
            list_jr.append(item_)
        return list_jr

    def parse_list_soup(self):
        start_time = time.time()
        self.list_item = []
        url = f'https://www.journeyinlife.net/search/label/phrase?v=full&page={self.page_number}'
        self.driver.get(url)
        print(f"url: {url}")
        time.sleep(time_loading)
        page_source = self.driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')

        row = soup.find('div', class_='row')
        list_phrase = row.find_all('div', class_='col-md-4')

        for i in list_phrase:
            item_journey = JourneyItem()
            item_journey.link = i.find('a').get('href')
            item_journey.thumb = i.find('a').find('img').get('data-src')
            title = i.find('div', class_='title').text

            item_journey.key = hashlib.md5(title.encode()).hexdigest()
            item_journey.title = title
            self.list_item.append(item_journey)

        # for detail in self.list_item:
        #     self.parse_detail(detail)

        self.writer_csv(list_item=self.list_item)
        print("--- %s seconds ---" % (time.time() - start_time))
        if self.page_number < self.page_total:
            self.page_number += 1
            self.parse_list_soup()

    # def tag_content(self, tag):
    #     return tag['class'] == 'col-md-4' and tag.parent['class'] == 'row'

    def parse_detail(self, item_journey):
        self.driver.get(item_journey.link)
        print(f"url: {item_journey.link}")
        page_source = self.driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')
        page = soup.find('div', class_='entry-body')
        item_journey.image = page.find('img').get('src')
        texts = page.findAll('div', {'style': 'text-align: justify;'})
        for text in texts:
            item_journey.str_content = f'{item_journey.str_content} \n ' \
                                       f'{text.text}'
            # item_journey.content.append(text.text)

        print(item_journey.title)

    def parse_detail_url(self, url: str, row_number: int):
        self.driver.get(url)
        print(f"url: {url}")
        page_source = self.driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')
        page = soup.find('div', class_='entry-body')
        if page is None:
            return
        image = page.find('img').get('src')
        texts = page.findAll('div', {'style': 'text-align: justify;'})
        str_content = ''
        for text in texts:
            str_content = f'{str_content} \n ' \
                          f'{text.text}'

        self.worksheet[f'D{row_number}'].value = str_content
        self.worksheet[f'F{row_number}'].value = image
        self.__save_ex()

    def parse_detail_jr(self, jr: JourneyItem) -> JourneyItem:
        self.driver.get(jr.link)
        print(f"url: {jr.link}")
        page_source = self.driver.page_source
        soup = BeautifulSoup(page_source, 'lxml')
        page = soup.find('div', class_='entry-body')
        if page is None:
            jr.content = "content_empty"
            return jr
        img = page.find('img')
        if img is not None:
            image = img.get('src')
            jr.image = image
        texts = page.findAll('div', {'style': 'text-align: justify;'})
        str_content = ''
        for text in texts:
            print(text.text)
            str_content = f'{str_content} \n ' \
                          f'{text.text}'
        if str_content == '':
            str_content = 'content_empty'
        jr.str_content = str_content
        return jr

    def close(self):
        self.workbook.save(f'journey_in_life_{self.page_total}.xlsx')
        self.workbook.close()
        if self.driver is not None:
            self.driver.close()
            self.driver.quit()

    def writer_csv(self, list_item):
        print('writer excel')
        self.init_excel()
        for i in range(len(list_item)):
            row_number = self.count + 1
            self.worksheet[f'A{row_number}'].value = f'{self.count}'
            self.worksheet[f'B{row_number}'].value = list_item[i].key
            self.worksheet[f'C{row_number}'].value = list_item[i].title
            self.worksheet[f'D{row_number}'].value = list_item[i].str_content
            self.worksheet[f'E{row_number}'].value = list_item[i].link
            self.worksheet[f'F{row_number}'].value = list_item[i].image
            self.worksheet[f'G{row_number}'].value = list_item[i].thumb
            self.count += 1
        # self.workbook.save(f'journey_in_life_{self.page_total}.xlsx')
        self.__save_ex()

    def get_db_and_scrawl(self):
        db = db_cassandra.db_cassandra()
        loop = True
        while loop:
            rows = db.get_db()
            if rows is None:
                loop = False
            else:
                for row in rows:
                    jr = self.__parse_from_db(row)
                    db.insert_db([jr])

    # INSERT INTO JOURNEY (key, title, link, thumb, content, image_url)
    def __parse_from_db(self, row) -> JourneyItem:
        jr = JourneyItem()
        jr.key = row.key
        jr.title = row.title
        jr.link = row.link
        jr.thumb = row.thumb
        return self.parse_detail_jr(jr)


if __name__ == '__main__':

    parse = JourneyParse()
    parse.init_selenium()
    # parse.init_excel()
    # list_jn = parse.get_data_excel_to_list()

    parse.get_db_and_scrawl()
    # db = db_cassandra.db_cassandra()
    # db.create_table_phrase()
    # db.insert_db(list_jn)
    # db.get_db()

    # parse.get_data_excel()
    # try:
    #     parse.parse_list_soup()
    # except:
    #     # parse.close()
    #     print("Caught it!")
    # finally:
    parse.close()
